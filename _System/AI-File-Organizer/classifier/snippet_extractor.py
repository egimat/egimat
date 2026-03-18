"""Extract text snippets from files for AI classification."""

import csv
import io
import os


def extract_snippet(filepath: str, max_words: int = 200) -> str:
    """Extract a text snippet from a file for classification.

    Supports: .txt, .md, .pdf (pdfplumber), .docx (python-docx),
              .csv, .xlsx (openpyxl).
    Returns empty string for binary/unknown types. Never crashes.
    """
    try:
        ext = os.path.splitext(filepath)[1].lower()

        if ext in (".txt", ".md"):
            return _extract_text(filepath, max_words)
        elif ext == ".pdf":
            return _extract_pdf(filepath, max_words)
        elif ext == ".docx":
            return _extract_docx(filepath, max_words)
        elif ext == ".csv":
            return _extract_csv(filepath, max_words)
        elif ext == ".xlsx":
            return _extract_xlsx(filepath, max_words)
        else:
            return ""
    except Exception:
        return ""


def _truncate(text: str, max_words: int) -> str:
    words = text.split()
    return " ".join(words[:max_words])


def _extract_text(filepath: str, max_words: int) -> str:
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        return _truncate(f.read(), max_words)


def _extract_pdf(filepath: str, max_words: int) -> str:
    import pdfplumber

    text_parts = []
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text_parts.append(page_text)
            if len(" ".join(text_parts).split()) >= max_words:
                break
    return _truncate(" ".join(text_parts), max_words)


def _extract_docx(filepath: str, max_words: int) -> str:
    import docx

    doc = docx.Document(filepath)
    text_parts = []
    for para in doc.paragraphs:
        if para.text.strip():
            text_parts.append(para.text)
        if len(" ".join(text_parts).split()) >= max_words:
            break
    return _truncate(" ".join(text_parts), max_words)


def _extract_csv(filepath: str, max_words: int) -> str:
    text_parts = []
    with open(filepath, "r", encoding="utf-8", errors="replace") as f:
        reader = csv.reader(f)
        for row in reader:
            text_parts.append(" ".join(row))
            if len(" ".join(text_parts).split()) >= max_words:
                break
    return _truncate(" ".join(text_parts), max_words)


def _extract_xlsx(filepath: str, max_words: int) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(filepath, read_only=True, data_only=True)
    text_parts = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                text_parts.append(" ".join(cells))
            if len(" ".join(text_parts).split()) >= max_words:
                break
        if len(" ".join(text_parts).split()) >= max_words:
            break
    wb.close()
    return _truncate(" ".join(text_parts), max_words)
